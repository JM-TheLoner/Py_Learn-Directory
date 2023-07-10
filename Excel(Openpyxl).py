import openpyxl as xl
from openpyxl.styles import Font


Master_workbook = xl.load_workbook("excel files\master_sheet.xlsx")   # type: ignore
Master_worksheet = Master_workbook['data']

daily_workbook = xl.load_workbook("excel files\daily_sheet.xlsx")  # type: ignore
daily_worksheet = daily_workbook['data']


for i in range(2, daily_worksheet.max_row+1):
    i_d = daily_worksheet.cell(row=i, column=1)
    t_p = daily_worksheet.cell(row=i, column=2)
    t_r = daily_worksheet.cell(row=i, column=3)

    id = i_d.value + 1 # type: ignore
    tp = t_p.value
    tr = t_r.value

    new = Master_worksheet.cell(row=id, column=6).value + tp  # type: ignore
    Master_worksheet.cell(row=id, column=6).value = new

    new2 = Master_worksheet.cell(row=id, column=7).value + tr  # type: ignore
    Master_worksheet.cell(row=id, column=7).value = new2

Master_workbook.save('excel files\master_sheet_1.xlsx')  # type: ignore

new_daily = xl.Workbook()
ws = new_daily.active

for i in range(2, Master_worksheet.max_column+1):
    ws.cell(row=1, column=i-1).value = Master_worksheet.cell(row=1, column=i).value  # type: ignore
    fonts = Font(name="New times roman", size=12, bold=True)
    ws.cell(row=1, column=i-1).font = fonts  # type: ignore


ids = []

for i in range(2, daily_worksheet.max_row+1):
    a = daily_worksheet.cell(row=i, column=1).value
    ids.append(a)


list = []
for rows in range (2, Master_worksheet.max_row+1):
    id = Master_worksheet.cell(row=rows, column=1).value
    if id in ids:
        lis = []
        for col in range (2, Master_worksheet.max_column+1):
            lis.append(Master_worksheet.cell(row=rows, column=col).value)
        list.append(lis)
   
for data in list:
    ws.append(data)  # type: ignore

new_daily.save("new_daily.xlsx")