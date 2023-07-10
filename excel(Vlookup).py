import openpyxl as xl

#v-lookup
# Master_workbook = xl.load_workbook("Py_Learn Directory\master_sheet.xlsx") 
# Master_worksheet = Master_workbook['data']

# daily_workbook = xl.load_workbook("Py_Learn Directory\daily_sheet.xlsx")
# daily_worksheet = daily_workbook['data']


# for i in daily_worksheet.iter_rows():
#     id = i[0].value
#     row_number = i[0].row
#     for n in Master_worksheet.iter_rows():
#         if n[0].value == id:
#             daily_worksheet.cell(row=row_number, column=4).value = n[1].value 
#             daily_worksheet.cell(row=row_number, column=5).value = n[2].value
#             daily_worksheet.cell(row=row_number, column=6).value = n[3].value


# daily_workbook.save('Py_Learn Directory\daily_sheet.xlsx')

name_data = xl.load_workbook("Py_Learn Directory\master_sheet.xlsx") 
names = name_data['Sheet1']

is_data = True
row_count = 1

while is_data:
    row_count+=1
    first_name = names.cell(row=row_count, column=1).value
    
    if first_name != None:
        names.cell(row=row_count, column=1).value = first_name.strip('"')
    else:
        is_data = False

name_data.save("Py_Learn Directory\master_sheet.xlsx")